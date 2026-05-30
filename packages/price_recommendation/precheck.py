"""Fast deterministic checks for price recommendation inputs."""

from __future__ import annotations

import csv
import io
import json
import os
import re
from collections import defaultdict
from statistics import mean, median
from typing import Any


PRODUCT_FIELD_KEYWORDS = (
    "商品", "品名", "名称", "药品", "sku", "product", "item", "条码", "货号"
)
PRICE_FIELD_KEYWORDS = (
    "售价", "价格", "单价", "零售价", "price", "sale_price", "retail"
)
SALES_FIELD_KEYWORDS = (
    "销量", "销售数量", "销售量", "数量", "件数", "订单量", "sales", "qty", "quantity", "销售额", "revenue"
)
TIME_FIELD_KEYWORDS = (
    "日期", "时间", "月份", "date", "time", "day", "month"
)
STORE_FIELD_KEYWORDS = (
    "门店", "店名", "店铺", "分店", "药店", "商户", "店号", "store", "shop", "branch", "pharmacy"
)


def run_precheck(decoded_files: list[dict], product_name: str) -> dict:
    """Return a user-facing validation result for uploaded files and product name."""
    product_name = (product_name or "").strip()
    issues: list[dict] = []
    warnings: list[dict] = []
    text_source_present = any(str(item.get("name") or "").lower().endswith(".txt") for item in decoded_files)

    if not product_name:
        issues.append({"code": "missing_product_name", "message": "缺少商品名称"})

    inspection = inspect_uploaded_files(decoded_files)
    tables = inspection["tables"]
    if inspection["errors"]:
        for err in inspection["errors"]:
            warnings.append({"code": "file_partial_parse_failed", "message": err})

    total_rows = sum(len(t["rows"]) for t in tables)
    fields = _collect_fields(tables)
    product_fields = _match_fields(fields, PRODUCT_FIELD_KEYWORDS)
    price_fields = _match_fields(fields, PRICE_FIELD_KEYWORDS)
    sales_fields = _match_fields(fields, SALES_FIELD_KEYWORDS)
    time_fields = _match_fields(fields, TIME_FIELD_KEYWORDS)
    matched_rows = _find_product_rows(tables, product_name, product_fields)
    matched_row_count = len(matched_rows)

    if not tables or total_rows == 0:
        if text_source_present:
            warnings.append({"code": "text_only_input", "message": "文本文件未解析出表格结构，将继续按内容尝试匹配"})
        else:
            issues.append({"code": "no_tabular_data", "message": "上传文件中没有可用于分析的表格数据"})
    if product_name and len(_normalize_text(product_name)) < 2:
        issues.append({"code": "product_name_too_short", "message": "商品名称过短，请提供更具体的完整品名"})
    elif product_name and len(_normalize_text(product_name)) < 4:
        warnings.append({"code": "product_name_may_be_broad", "message": "商品名称可能偏宽，建议补充规格或完整品名"})
    if tables and not product_fields:
        if text_source_present:
            warnings.append({"code": "product_field_missing", "message": "文本中未识别到明确的商品字段，将按全文内容继续匹配"})
        else:
            issues.append({"code": "product_field_missing", "message": "未识别到商品名称、品名或 SKU 字段"})
    if tables and not price_fields:
        if text_source_present:
            warnings.append({"code": "price_field_missing", "message": "文本中未识别到明确的价格字段，将尝试从价格语句中提取数字"})
        else:
            issues.append({"code": "price_field_missing", "message": "未识别到价格、售价或单价字段"})
    if tables and not sales_fields:
        if text_source_present:
            warnings.append({"code": "sales_field_missing", "message": "文本中未识别到明确的销量字段，将继续尝试从内容中提取"})
        else:
            issues.append({"code": "sales_field_missing", "message": "未识别到销量、销售额或订单量字段"})
    if product_name and product_fields and matched_row_count == 0:
        if text_source_present:
            warnings.append({"code": "product_not_found", "message": "文本中未找到明确的商品记录，将继续允许后续流程处理"})
        else:
            issues.append({"code": "product_not_found", "message": "上传文件中没有找到该商品的销售记录"})
    if tables and not time_fields:
        warnings.append({"code": "time_field_missing", "message": "未识别到日期或时间字段，后续趋势判断会受限"})

    valid = not issues
    match_confidence = _confidence(
        matched_row_count,
        total_rows,
        bool(price_fields),
        bool(sales_fields),
        bool(time_fields),
        bool(product_fields),
        product_name,
    )
    if text_source_present:
        match_confidence = min(match_confidence, 0.72 if matched_row_count else 0.55)

    return {
        "status": "ok" if valid else "failed",
        "valid": valid,
        "productName": product_name,
        "productNameQuality": {
            "status": "ok" if product_name and len(_normalize_text(product_name)) >= 4 else ("warning" if product_name else "failed"),
            "message": "商品名称足够具体" if product_name and len(_normalize_text(product_name)) >= 4 else "商品名称需要更具体",
        },
        "fileQuality": {
            "status": "ok" if tables and total_rows > 0 else ("warning" if text_source_present else "failed"),
            "message": f"文件可解析，识别到 {len(tables)} 张表、{total_rows} 行数据" if tables else ("文本文件已接收，正在按内容匹配" if text_source_present else "文件未解析出表格数据"),
        },
        "productDataMatch": {
            "status": "ok" if matched_row_count > 0 else ("warning" if text_source_present else "failed"),
            "matchedRows": matched_row_count,
            "confidence": match_confidence,
        },
        "detectedFields": {
            "productFields": product_fields,
            "priceFields": price_fields,
            "salesFields": sales_fields,
            "timeFields": time_fields,
        },
        "issues": issues,
        "warnings": warnings,
    }


def inspect_uploaded_files(decoded_files: list[dict], max_rows_per_table: int = 1000) -> dict:
    """Parse JSON/CSV/XLSX uploads into lightweight table dictionaries."""
    tables: list[dict] = []
    errors: list[str] = []
    for item in decoded_files:
        name = item.get("name") or "unnamed"
        data = item.get("bytes") or b""
        lower = name.lower()
        try:
            if lower.endswith(".json"):
                raw = json.loads(data.decode("utf-8-sig"))
                tables.extend(_json_to_tables(raw, os.path.splitext(name)[0], max_rows_per_table))
            elif lower.endswith(".csv"):
                tables.append(_csv_to_table(data, name, max_rows_per_table))
            elif lower.endswith(".txt"):
                tables.extend(_txt_to_tables(data, name, max_rows_per_table))
            elif lower.endswith((".xlsx", ".xls")):
                tables.extend(_xlsx_to_tables(data, name, max_rows_per_table))
        except Exception as exc:
            errors.append(f"{name} 解析失败: {exc}")
    return {"tables": tables, "errors": errors}


def build_basic_recommendation(inspection: dict, product_name: str, candidate_count: int = 2) -> dict:
    """Build a deterministic baseline recommendation for the first implementation."""
    tables = inspection.get("tables", [])
    fields = _collect_fields(tables)
    product_fields = _match_fields(fields, PRODUCT_FIELD_KEYWORDS)
    price_fields = _match_fields(fields, PRICE_FIELD_KEYWORDS)
    sales_fields = _match_fields(fields, SALES_FIELD_KEYWORDS)
    time_fields = _match_fields(fields, TIME_FIELD_KEYWORDS)
    matched_rows = _find_product_rows(tables, product_name, product_fields)
    price_field = price_fields[0] if price_fields else ""

    price_values: list[float] = []
    for row in matched_rows:
        price_values.extend(_extract_price_candidates_from_row(row, price_field))

    if not price_values:
        raise ValueError("目标商品没有可用的历史价格数据")

    median_price = round(median(price_values), 2)
    mean_price = round(mean(price_values), 2)
    candidates = [median_price]
    if abs(mean_price - median_price) >= 0.01:
        candidates.append(mean_price)
    else:
        candidates.append(round(median_price * 1.03, 2))

    recommendations = []
    for idx, price in enumerate(_dedupe_prices(candidates)[:max(1, candidate_count)], start=1):
        recommendations.append({
            "rank": idx,
            "price": price,
            "unit": "元",
            "reason": "基于目标商品历史价格分布生成的基准推荐，后续可替换为价格弹性或曲线拟合模型",
            "confidence": _confidence(
                len(matched_rows),
                sum(len(t["rows"]) for t in tables),
                bool(price_values),
                bool(sales_fields),
                bool(time_fields),
                bool(product_fields),
                product_name,
            ),
        })

    return {
        "taskType": "price_recommendation",
        "productName": product_name,
        "recommendations": recommendations,
        "validPriceRange": {
            "min": round(min(price_values), 2),
            "max": round(max(price_values), 2),
            "unit": "元",
        },
        "evidence": {
            "matchedRows": len(matched_rows),
            "observedPriceCount": len(price_values),
            "priceField": price_field,
            "salesField": sales_fields[0] if sales_fields else "",
            "timeField": time_fields[0] if time_fields else "",
            "sourceTables": [t["name"] for t in tables],
            "notes": ["当前为确定性基准推荐，尚未接入价格弹性模型"],
        },
        "warnings": [],
    }


def build_price_point_artifacts(inspection: dict, product_name: str) -> dict:
    """Build raw and normalized price points for the first price workflow."""
    tables = inspection.get("tables", [])
    fields = _collect_fields(tables)
    product_fields = _match_fields(fields, PRODUCT_FIELD_KEYWORDS)
    price_fields = _match_fields(fields, PRICE_FIELD_KEYWORDS)
    sales_fields = _match_fields(fields, SALES_FIELD_KEYWORDS)
    time_fields = _match_fields(fields, TIME_FIELD_KEYWORDS)
    store_fields = _match_fields(fields, STORE_FIELD_KEYWORDS)
    matched_entries = _find_product_entries(tables, product_name, product_fields)

    raw_points: list[dict[str, Any]] = []
    for entry in matched_entries:
        row = entry["row"]
        price_values = _extract_price_candidates_from_row(row, price_fields[0] if price_fields else "")
        if not price_values:
            continue
        raw_qty = _extract_sales_value(row, sales_fields)
        source_shop = _extract_store_value(row, store_fields) or entry["table"]
        source_date = _extract_time_value(row, time_fields)
        for price in price_values:
            raw_points.append({
                "price": round(price, 2),
                "rawQty": round(raw_qty, 4),
                "sourceShop": source_shop,
                "sourceTable": entry["table"],
                "sourceDate": source_date,
            })

    if not raw_points:
        raise ValueError("目标商品没有可用的价格点数据")

    normalized = _normalize_price_points(raw_points)
    return {
        "productName": product_name,
        "raw": {
            "productName": product_name,
            "points": raw_points,
            "fields": {
                "productField": product_fields[0] if product_fields else "",
                "priceField": price_fields[0] if price_fields else "",
                "salesField": sales_fields[0] if sales_fields else "",
                "timeField": time_fields[0] if time_fields else "",
                "storeField": store_fields[0] if store_fields else "",
            },
        },
        "normalized": normalized,
        "evidence": {
            "matchedRows": len(matched_entries),
            "rawPointCount": len(raw_points),
            "priceField": price_fields[0] if price_fields else "",
            "salesField": sales_fields[0] if sales_fields else "",
            "timeField": time_fields[0] if time_fields else "",
            "storeField": store_fields[0] if store_fields else "",
            "sourceTables": [t["name"] for t in tables],
        },
    }


def build_recommendation_from_points(
    *,
    product_name: str,
    normalized_points: list[dict[str, Any]],
    evidence: dict[str, Any],
    candidate_count: int = 2,
) -> dict:
    """Build the first recommendation payload from normalized discrete points."""
    if not normalized_points:
        raise ValueError("缺少归一化价格点")

    scored_points = []
    for point in normalized_points:
        price = _to_number(point.get("price"))
        qty = _to_number(point.get("normalizedQty"))
        if price is None or qty is None:
            continue
        scored_points.append({
            "price": round(price, 2),
            "normalizedQty": round(qty, 4),
            "score": round(price * qty, 4),
            "sampleCount": int(_to_number(point.get("sampleCount")) or 0),
        })

    if not scored_points:
        raise ValueError("归一化价格点缺少有效数值")

    scored_points.sort(key=lambda item: (-item["score"], item["price"]))
    recommendations = []
    total_points = max(len(scored_points), 1)
    for rank, item in enumerate(scored_points[:max(1, candidate_count)], start=1):
        recommendations.append({
            "rank": rank,
            "price": item["price"],
            "unit": "元",
            "reason": f"归一化后销售额最高，折算销量 {item['normalizedQty']:.2f}",
            "confidence": _confidence(
                evidence.get("matchedRows", 0),
                max(evidence.get("rawPointCount", 0), total_points),
                bool(evidence.get("priceField")),
                bool(evidence.get("salesField")),
                bool(evidence.get("timeField")),
                True,
                product_name,
            ),
        })

    prices = [item["price"] for item in scored_points]
    result_evidence = dict(evidence)
    result_evidence["observedPriceCount"] = len(scored_points)
    result_evidence["notes"] = ["当前结果直接使用归一化后的离散点集生成"]
    return {
        "taskType": "price_recommendation",
        "productName": product_name,
        "recommendations": recommendations,
        "normalizedPoints": normalized_points,
        "validPriceRange": {
            "min": round(min(prices), 2),
            "max": round(max(prices), 2),
            "unit": "元",
        },
        "evidence": result_evidence,
        "warnings": [],
    }


def _json_to_tables(raw: Any, name: str, max_rows: int) -> list[dict]:
    tables: list[dict] = []
    if isinstance(raw, list):
        return [{"name": name, "rows": _dict_rows(raw[:max_rows])}]
    if isinstance(raw, dict):
        for key, value in raw.items():
            if isinstance(value, list) and value:
                tables.append({"name": f"{name}/{key}", "rows": _dict_rows(value[:max_rows])})
            elif isinstance(value, dict):
                for sub_key, sub_value in value.items():
                    if isinstance(sub_value, list) and sub_value:
                        tables.append({"name": f"{name}/{key}/{sub_key}", "rows": _dict_rows(sub_value[:max_rows])})
        if not tables:
            tables.append({"name": name, "rows": [raw]})
    return tables


def _csv_to_table(data: bytes, name: str, max_rows: int) -> dict:
    content = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    return {"name": name, "rows": [dict(row) for _, row in zip(range(max_rows), reader)]}


def _xlsx_to_tables(data: bytes, name: str, max_rows: int) -> list[dict]:
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    tables: list[dict] = []
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows_iter = sheet.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if not header:
                continue
            columns = [str(value).strip() if value is not None else f"col_{idx}" for idx, value in enumerate(header)]
            rows = []
            for idx, values in enumerate(rows_iter):
                if idx >= max_rows:
                    break
                if not any(value is not None for value in values):
                    continue
                rows.append(dict(zip(columns, values)))
            tables.append({"name": f"{name}/{sheet_name}", "rows": rows})
    finally:
        workbook.close()
    return tables


def _txt_to_tables(data: bytes, name: str, max_rows: int) -> list[dict]:
    text = _decode_text(data)
    lines = [line.strip() for line in text.splitlines()]
    rows: list[dict] = []
    for idx, line in enumerate(lines, start=1):
        if not line:
            continue
        row: dict[str, Any] = {"line_no": idx, "content": line}
        row.update(_extract_text_row_fields(line))
        rows.append(row)
        if len(rows) >= max_rows:
            break
    if not rows and text.strip():
        rows.append({"line_no": 1, "content": text.strip()})
    return [{"name": name, "rows": rows}]


def _dict_rows(values: list[Any]) -> list[dict]:
    rows: list[dict] = []
    for value in values:
        if isinstance(value, dict):
            rows.append(value)
        else:
            rows.append({"value": value})
    return rows


def _collect_fields(tables: list[dict]) -> list[str]:
    fields: set[str] = set()
    for table in tables:
        for row in table.get("rows", []):
            if isinstance(row, dict):
                fields.update(str(key) for key in row.keys())
    return sorted(fields)


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            return data.decode(encoding)
        except Exception:
            continue
    return data.decode("utf-8", errors="replace")


def _extract_text_row_fields(line: str) -> dict[str, str]:
    row: dict[str, str] = {}
    parts = re.split(r"[,\t|;；]+", line)
    for part in parts:
        segment = part.strip()
        if not segment:
            continue
        if any(sep in segment for sep in (":", "：", "=")):
            key, value = re.split(r"[:：=]", segment, maxsplit=1)
            mapped = _map_text_field_name(key)
            if mapped and value.strip():
                row[mapped] = value.strip()
            elif key.strip() and value.strip():
                row[key.strip()] = value.strip()

    if "商品名称" not in row:
        value = _extract_text_value(line, PRODUCT_FIELD_KEYWORDS)
        if value:
            row["商品名称"] = value
    if "价格" not in row:
        value = _extract_numeric_text_value(line, PRICE_FIELD_KEYWORDS)
        if value:
            row["价格"] = value
    if "销量" not in row:
        value = _extract_numeric_text_value(line, SALES_FIELD_KEYWORDS)
        if value:
            row["销量"] = value
    if "日期" not in row:
        value = _extract_text_value(line, TIME_FIELD_KEYWORDS)
        if value:
            row["日期"] = value
    return row


def _map_text_field_name(key: str) -> str | None:
    normalized = _normalize_text(key)
    if any(keyword in normalized for keyword in PRODUCT_FIELD_KEYWORDS):
        return "商品名称"
    if any(keyword in normalized for keyword in PRICE_FIELD_KEYWORDS):
        return "价格"
    if any(keyword in normalized for keyword in SALES_FIELD_KEYWORDS):
        return "销量"
    if any(keyword in normalized for keyword in TIME_FIELD_KEYWORDS):
        return "日期"
    return None


def _extract_text_value(line: str, keywords: tuple[str, ...]) -> str:
    label_pattern = "|".join(re.escape(keyword) for keyword in keywords)
    patterns = [
        rf"(?:{label_pattern})\s*[:：=]?\s*([^,，;；|\t]+)",
        rf"(?:{label_pattern})\s*(?:是|为|：|:)?\s*([^,，;；|\t]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, line, flags=re.IGNORECASE)
        if match:
            value = match.group(1).strip()
            if value:
                return value
    return ""


def _extract_numeric_text_value(line: str, keywords: tuple[str, ...]) -> str:
    label_pattern = "|".join(re.escape(keyword) for keyword in keywords)
    patterns = [
        rf"(?:{label_pattern})\s*[:：=]?\s*(?:￥|¥|RMB|元)?\s*(-?\d+(?:\.\d+)?)",
        rf"(?:{label_pattern})\s*(?:是|为|：|:)?\s*(?:￥|¥|RMB|元)?\s*(-?\d+(?:\.\d+)?)",
    ]
    for pattern in patterns:
        match = re.search(pattern, line, flags=re.IGNORECASE)
        if match:
            return match.group(1)
    return ""


def _match_fields(fields: list[str], keywords: tuple[str, ...]) -> list[str]:
    matched = []
    for field in fields:
        lowered = field.lower()
        if any(keyword.lower() in lowered for keyword in keywords):
            matched.append(field)
    return matched


def _find_product_rows(tables: list[dict], product_name: str, product_fields: list[str]) -> list[dict]:
    target = _normalize_text(product_name)
    if not target:
        return []
    rows = []
    search_fields = product_fields or _collect_fields(tables)
    for table in tables:
        for row in table.get("rows", []):
            if not isinstance(row, dict):
                continue
            for field in search_fields:
                value = row.get(field)
                text = _normalize_text(value)
                if text and (target in text or text in target):
                    rows.append(row)
                    break
    return rows


def _find_product_entries(tables: list[dict], product_name: str, product_fields: list[str]) -> list[dict]:
    target = _normalize_text(product_name)
    if not target:
        return []
    entries = []
    search_fields = product_fields or _collect_fields(tables)
    for table in tables:
        table_name = table.get("name", "")
        for row in table.get("rows", []):
            if not isinstance(row, dict):
                continue
            for field in search_fields:
                value = row.get(field)
                text = _normalize_text(value)
                if text and (target in text or text in target):
                    entries.append({"table": table_name, "row": row})
                    break
    return entries


def _normalize_text(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()


def _to_number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    text = str(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _confidence(
    matched_rows: int,
    total_rows: int,
    has_price: bool,
    has_sales: bool,
    has_time: bool = False,
    has_product: bool = False,
    product_name: str = "",
) -> float:
    score = 0.2
    normalized_name = _normalize_text(product_name)
    if normalized_name:
        if len(normalized_name) >= 12:
            score += 0.16
        elif len(normalized_name) >= 8:
            score += 0.12
        elif len(normalized_name) >= 4:
            score += 0.08
        else:
            score += 0.02
    if total_rows > 0 and matched_rows:
        coverage = matched_rows / max(total_rows, 1)
        score += min(0.24, coverage * 0.35)
    elif total_rows > 0:
        score += 0.02
    if has_product:
        score += 0.08
    if has_price:
        score += 0.16
    if has_sales:
        score += 0.1
    if has_time:
        score += 0.05
    return round(min(max(score, 0.05), 0.95), 2)


def _dedupe_prices(values: list[float]) -> list[float]:
    result: list[float] = []
    seen: set[float] = set()
    for value in values:
        rounded = round(value, 2)
        if rounded not in seen and rounded > 0:
            result.append(rounded)
            seen.add(rounded)
    return result


def _extract_price_candidates_from_row(row: dict, price_field: str) -> list[float]:
    candidates: list[float] = []
    if price_field:
        value = _to_number(row.get(price_field))
        if value is not None:
            candidates.append(value)
    for key in ("价格", "售价", "单价", "零售价", "retail", "sale_price", "price"):
        value = _to_number(row.get(key))
        if value is not None:
            candidates.append(value)
    content = str(row.get("content") or "")
    if content:
        candidates.extend(_extract_price_numbers_from_text(content))
    if not candidates:
        for key, value in row.items():
            if key == "line_no":
                continue
            if isinstance(value, (int, float)):
                number = float(value)
                if number > 0:
                    candidates.append(number)
            elif isinstance(value, str):
                candidates.extend(_extract_price_numbers_from_text(value))
    return _dedupe_prices(candidates)


def _extract_price_numbers_from_text(text: str) -> list[float]:
    if not text:
        return []
    price_matches: list[float] = []
    keyword_pattern = r"(?:售价|价格|单价|零售价|price|sale_price|retail)"
    labeled_patterns = [
        rf"{keyword_pattern}\s*[:：=]?\s*(?:￥|¥|RMB|元)?\s*(-?\d+(?:\.\d+)?)",
        rf"(?:￥|¥|RMB|元)\s*(-?\d+(?:\.\d+)?)",
    ]
    for pattern in labeled_patterns:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            number = _to_number(match.group(1))
            if number is not None:
                price_matches.append(number)
    if price_matches:
        return price_matches
    generic_numbers: list[float] = []
    for match in re.finditer(r"-?\d+(?:\.\d+)?", text):
        number = _to_number(match.group(0))
        if number is not None and number > 0:
            generic_numbers.append(number)
    return generic_numbers


def _extract_sales_value(row: dict, sales_fields: list[str]) -> float:
    for field in sales_fields:
        value = _to_number(row.get(field))
        if value is not None and value > 0:
            return value
    for key in ("销量", "销售数量", "销售量", "数量", "件数", "订单量", "qty", "quantity", "sales"):
        value = _to_number(row.get(key))
        if value is not None and value > 0:
            return value
    content = str(row.get("content") or "")
    if content:
        value = _to_number(_extract_numeric_text_value(content, SALES_FIELD_KEYWORDS))
        if value is not None and value > 0:
            return value
    return 1.0


def _extract_store_value(row: dict, store_fields: list[str]) -> str:
    for field in store_fields:
        value = str(row.get(field) or "").strip()
        if value:
            return value
    for key in ("门店", "门店名称", "店名", "店铺", "药店", "shop", "store", "pharmacy"):
        value = str(row.get(key) or "").strip()
        if value:
            return value
    return ""


def _extract_time_value(row: dict, time_fields: list[str]) -> str:
    for field in time_fields:
        value = str(row.get(field) or "").strip()
        if value:
            return value
    for key in ("日期", "时间", "月份", "date", "time", "day", "month"):
        value = str(row.get(key) or "").strip()
        if value:
            return value
    return ""


def _normalize_price_points(raw_points: list[dict[str, Any]]) -> dict[str, Any]:
    shop_totals: dict[str, float] = defaultdict(float)
    for point in raw_points:
        shop = str(point.get("sourceShop") or "unknown")
        qty = _to_number(point.get("rawQty")) or 0.0
        shop_totals[shop] += max(qty, 0.0)

    totals = [value for value in shop_totals.values() if value > 0]
    baseline_qty = median(totals) if totals else 1.0
    price_groups: dict[float, dict[str, Any]] = {}

    for point in raw_points:
        shop = str(point.get("sourceShop") or "unknown")
        price = round(_to_number(point.get("price")) or 0.0, 2)
        raw_qty = _to_number(point.get("rawQty")) or 0.0
        shop_total = shop_totals.get(shop) or baseline_qty or 1.0
        factor = baseline_qty / shop_total if shop_total > 0 else 1.0
        normalized_qty = raw_qty * factor

        bucket = price_groups.setdefault(price, {
            "price": price,
            "rawQty": 0.0,
            "normalizedQty": 0.0,
            "sampleCount": 0,
            "sourceShops": set(),
            "factorSum": 0.0,
        })
        bucket["rawQty"] += raw_qty
        bucket["normalizedQty"] += normalized_qty
        bucket["sampleCount"] += 1
        bucket["sourceShops"].add(shop)
        bucket["factorSum"] += factor

    points = []
    for price in sorted(price_groups):
        bucket = price_groups[price]
        sample_count = max(bucket["sampleCount"], 1)
        points.append({
            "price": round(bucket["price"], 2),
            "rawQty": round(bucket["rawQty"], 4),
            "normalizedQty": round(bucket["normalizedQty"], 4),
            "sampleCount": sample_count,
            "avgFactor": round(bucket["factorSum"] / sample_count, 6),
            "sourceShops": sorted(bucket["sourceShops"]),
        })

    return {
        "points": points,
        "normalization": {
            "method": "shop_scale_median",
            "baselineQty": round(float(baseline_qty), 4),
            "shopCount": len(shop_totals),
        },
    }
