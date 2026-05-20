"""
input_adapter.py — 统一输入适配器
支持 JSON（含药店ERP）/ Excel / CSV 多文件解析，输出 DatasetBundle
"""
import json
import io
import os
from pathlib import Path
from typing import List, Optional


def _is_pharmacy_json(raw: dict) -> bool:
    """检测是否为旧版药店ERP JSON（含 page.module 和 businessTable.rows）"""
    if not isinstance(raw, dict):
        return False
    page = raw.get("page", {})
    if not isinstance(page, dict) or "module" not in page:
        return False
    return True


def _extract_pharmacy_tables(raw: dict, name: str) -> list:
    """药店 ERP JSON → 提取 inner table(s) 为 RawTable"""
    tables = []
    module = raw.get("page", {}).get("module", "unknown")
    view_type = raw.get("page", {}).get("viewType", "")

    # business_overview / o2o_business_summary → 提取 businessTable.rows
    biz_table = raw.get("businessTable") or raw.get("dailyBusinessTable")
    if biz_table and isinstance(biz_table, dict) and "rows" in biz_table:
        rows = biz_table["rows"]
        if rows:
            tables.append({"name": f"{name}/business", "rows": rows})

    # 提取 sourceDistribution
    src_dist = raw.get("sourceDistribution")
    if src_dist and isinstance(src_dist, dict) and "items" in src_dist:
        items = src_dist["items"]
        if items:
            tables.append({"name": f"{name}/channels", "rows": items})

    # operation_hot_products → ranking
    ranking = raw.get("ranking")
    if ranking and isinstance(ranking, list) and ranking:
        tables.append({"name": f"{name}/ranking", "rows": ranking})

    # hot_sale_top500 → products
    products = raw.get("products")
    if products and isinstance(products, list) and products:
        tables.append({"name": f"{name}/top500", "rows": products})

    # summary.metrics
    summary = raw.get("summary")
    if summary and isinstance(summary, dict) and "metrics" in summary:
        tables.append({"name": f"{name}/summary", "rows": summary["metrics"]})

    if not tables:
        # 兜底：整条 JSON 当一行
        return [{"name": name, "rows": [{"_meta": f"module={module} view={view_type}"}]}]

    return tables


def _json_to_tables_recursive(raw, name: str = "data") -> list:
    """递归展平 JSON → RawTable[]

    从 JSON 的任意深度提取所有数组为独立表，
    路径作为表名（如 data/level1/items）。
    兜底：无数组时整条 JSON 包成一行。
    """
    tables = []

    if isinstance(raw, list) and raw:
        tables.append({"name": name, "rows": raw})

    def _walk(obj, path):
        if isinstance(obj, dict):
            for key, val in obj.items():
                child_path = f"{path}/{key}"
                if isinstance(val, list) and val:
                    tables.append({"name": child_path, "rows": val})
                elif isinstance(val, dict):
                    _walk(val, child_path)

    if isinstance(raw, dict):
        _walk(raw, name)

    if not tables:
        if isinstance(raw, dict):
            return [{"name": name, "rows": [raw]}]
        return [{"name": name, "rows": []}]

    return tables


def _json_to_table(raw, name: str = "data") -> dict:
    """JSON → RawTable：数组转表，对象保留层级路径"""
    if isinstance(raw, list):
        return {"name": name, "rows": raw}
    if isinstance(raw, dict):
        rows = []
        for key, val in raw.items():
            if isinstance(val, list):
                rows.extend(val)
        if rows:
            return {"name": name, "rows": rows}
        return {"name": name, "rows": [raw]}
    return {"name": name, "rows": []}


def _excel_to_tables(file_bytes: bytes, filename: str) -> list:
    """Excel → RawTable[]：每 sheet 一张表"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("需要 openpyxl 解析 Excel 文件，请执行: pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    tables = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue
        header = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(all_rows[0])]
        rows = [dict(zip(header, r)) for r in all_rows[1:] if any(v is not None for v in r)]
        tables.append({"name": f"{filename}/{sheet_name}", "rows": rows})
    wb.close()
    return tables


def _detect_key_value_table(rows: list) -> bool:
    """通用检测：是否 key-value 结构（有 key 列 + 数值 value 列 + key 值数量合理）"""
    if not rows:
        return False

    all_cols = set()
    for row in rows:
        if isinstance(row, dict):
            all_cols.update(row.keys())

    if "key" not in all_cols or "value" not in all_cols:
        return False

    key_vals = set()
    val_is_num = True
    for row in rows:
        if isinstance(row, dict):
            k = row.get("key")
            v = row.get("value")
            if k is not None:
                key_vals.add(str(k))
            if v is not None and not isinstance(v, (int, float)):
                val_is_num = False

    if not key_vals or len(key_vals) > 50:
        return False
    if not val_is_num:
        return False
    return True


def _pivot_table(table: dict) -> tuple:
    """将 key-value 表 pivot 为宽表，并生成维度表

    返回: (pivoted_table, dim_table_or_None)
    - pivoted: key 各值→列名, value→值 (一行的宽表)
    - dim:     保留非key/value的列（label, unit等），供AI识别上下文
    """
    rows = table.get("rows", [])
    pivoted = {}
    dim_rows = []

    for row in rows:
        if isinstance(row, dict):
            k = str(row.get("key", "")).strip()
            v = row.get("value")
            if k:
                pivoted[k] = v

            dim_row = {"dim_key": k} if k else {}
            for col, val in row.items():
                if col not in ("key", "value") and not isinstance(val, (dict, list)):
                    dim_row[f"dim_{col}"] = val
            if dim_row and len(dim_row) > 1:
                dim_rows.append(dim_row)

    if not pivoted:
        return (table, None)

    name = table.get("name", "pivoted")
    pivoted_table = {"name": name, "rows": [pivoted]}

    dim_table = None
    if dim_rows:
        dim_table = {"name": f"{name}_dim", "rows": dim_rows}

    return (pivoted_table, dim_table)


def _csv_to_table(file_bytes: bytes, filename: str) -> dict:
    """CSV → RawTable"""
    import csv
    content = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    rows = [dict(row) for row in reader]
    return {"name": filename, "rows": rows}


def parse_file(file_bytes: bytes, filename: str):
    """根据后缀分发解析"""
    lower = filename.lower()
    if lower.endswith(".json"):
        raw = json.loads(file_bytes.decode("utf-8-sig"))
        if _is_pharmacy_json(raw):
            return _extract_pharmacy_tables(raw, os.path.splitext(filename)[0])
        return _json_to_tables_recursive(raw, os.path.splitext(filename)[0])
    elif lower.endswith(".xlsx") or lower.endswith(".xls"):
        return _excel_to_tables(file_bytes, filename)
    elif lower.endswith(".csv"):
        return _csv_to_table(file_bytes, filename)
    else:
        raise ValueError(f"不支持的文件类型: {filename}")


def parse_uploaded_files(decoded_files: List[dict]) -> dict:
    """
    批量解析上传文件，输出 DatasetBundle

    输入: [{"name": "概览-日.json", "bytes": b"..."}, ...]
    输出: DatasetBundle = {
        "source_type": "json|excel|csv",
        "tables": [RawTable],
        "received_at": "2026-05-12T10:00:00"
    }
    """
    from datetime import datetime

    all_tables = []
    source_type = "json"  # 默认

    for item in decoded_files:
        name = item.get("name", "unnamed")
        data_bytes = item.get("bytes", b"")
        if not data_bytes:
            continue

        lower = name.lower()
        if lower.endswith(".xlsx") or lower.endswith(".xls"):
            source_type = "excel"
        elif lower.endswith(".csv"):
            source_type = "csv"

        result = parse_file(data_bytes, name)
        if isinstance(result, list):
            all_tables.extend(result)
        else:
            all_tables.append(result)

    # 后处理：检测 key-value 结构并 pivot
    pivoted_count = 0
    for i, table in enumerate(all_tables):
        rows = table.get("rows", [])
        if _detect_key_value_table(rows):
            pivoted, dim = _pivot_table(table)
            all_tables[i] = pivoted
            if "original_name" not in all_tables[i]:
                all_tables[i]["original_name"] = table.get("name", "")
            if dim:
                all_tables.append(dim)
            pivoted_count += 1

    return {
        "source_type": source_type,
        "tables": all_tables,
        "pivoted_count": pivoted_count,
        "received_at": datetime.now().isoformat()
    }


def infer_source_type(filename: str) -> str:
    """从文件名推断来源类型"""
    lower = filename.lower()
    if lower.endswith(".json"):
        return "json"
    elif lower.endswith(".xlsx") or lower.endswith(".xls"):
        return "excel"
    elif lower.endswith(".csv"):
        return "csv"
    return "unknown"
