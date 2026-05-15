"""Agent 路由：POST /api/agent/analyze?pipeline=smol|pydantic。

单路由，pipeline 参数切换 PydanticPipeline 或 SmolPipeline，带文件解析。"""

from typing import Optional
from fastapi import APIRouter, Header, HTTPException, UploadFile, File, Query
from packages.agents.models import DatasetBundle, RawTable
from packages.agents.pydantic_pipeline import PydanticPipeline
from packages.agents.smol_pipeline import SmolPipeline

router = APIRouter(prefix="/api/agent")


@router.post("/analyze")
async def agent_analyze(
    files: list[UploadFile] = File(...),
    pipeline: str = Query("smol"),
    x_fzt_key: Optional[str] = Header(default=None),
):
    """单路由，按 pipeline 参数选择管线"""
    tables = _parse_uploads(files)
    bundle = DatasetBundle(tables=tables)

    if pipeline == "pydantic":
        pipe = PydanticPipeline()
    else:
        pipe = SmolPipeline()

    result = await pipe.run(bundle)
    return result.model_dump()


def _parse_uploads(files: list[UploadFile]) -> list[RawTable]:
    tables = []
    for f in files:
        raw = f.file.read()
        filename = f.filename or "unnamed"
        lower = filename.lower()

        if lower.endswith(".json"):
            import json as _json
            parsed = _json.loads(raw.decode("utf-8-sig"))
            if isinstance(parsed, list):
                tables.append(RawTable(name=filename, rows=parsed))
            elif isinstance(parsed, dict):
                tables.append(RawTable(name=filename, rows=[parsed]))
            else:
                raise HTTPException(400, f"不支持的 JSON 格式: {filename}")

        elif lower.endswith(".csv"):
            import csv, io as _io
            content = raw.decode("utf-8-sig")
            reader = csv.DictReader(_io.StringIO(content))
            rows = [dict(r) for r in reader]
            tables.append(RawTable(name=filename, rows=rows))

        elif lower.endswith(".xlsx") or lower.endswith(".xls"):
            import openpyxl, io as _io
            wb = openpyxl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    continue
                header = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(all_rows[0])]
                rows = [dict(zip(header, r)) for r in all_rows[1:] if any(v is not None for v in r)]
                tables.append(RawTable(name=f"{filename}/{sheet_name}", rows=rows))
            wb.close()

        else:
            raise HTTPException(400, f"不支持的文件类型: {filename}")

    if not tables:
        raise HTTPException(400, "未解析出有效数据")
    return tables
